import csv
import os
import requests
import argparse
from bs4 import BeautifulSoup
from urllib.parse import urlparse
import sys
from tabulate import tabulate
from openpyxl import Workbook
import re
from pathlib import Path

# 设置控制台编码为 UTF-8
sys.stdout.reconfigure(encoding='utf-8')

def is_txt_file(file_path):
    return file_path.endswith('.txt')

def is_url_alive(url):
    try:
        response = requests.head(url)
        return response.status_code == 200
    except requests.ConnectionError:
        return False

def get_title(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser', from_encoding="utf-8")
    return soup.title.string if soup.title else ''

def sanitize_sheet_name(sheet_name):
    # 替换特殊字符为下划线
    sanitized_name = re.sub(r'[\/:*?"<>|\\\\]', '_', sheet_name)
    return sanitized_name[:31]  # 工作表名称最多支持31个字符


def classify_urls(txt_file_path):
    if not os.path.isfile(txt_file_path):
        raise ValueError('Invalid file path')

    if not is_txt_file(txt_file_path):
        raise ValueError('Invalid file type. Only TXT files are allowed.')

    # 创建工作簿
    wb = Workbook()

    # 创建等号工作表
    equal_ws = wb.create_sheet(title=sanitize_sheet_name('等号'))

    # 创建杂工作表
    misc_ws = wb.create_sheet(title=sanitize_sheet_name('杂'))

    # 创建子域名工作表
    subdomain_ws = wb.create_sheet(title=sanitize_sheet_name('子域名'))

    with open(txt_file_path, 'r', encoding='utf-8') as txt_file:
        urls = txt_file.readlines()

    domain_set = set()
    equal_urls = []
    misc_urls = []

    for url in urls:
        url = url.strip()

        if not url:
            continue

        if '=' in url:
            equal_urls.append(url)
            continue

        parsed_url = urlparse(url)

        if parsed_url.scheme and parsed_url.netloc:
            domain = parsed_url.netloc
            domain_set.add(domain)
            subdomain_ws.append([domain])
        else:
            misc_urls.append(url)
    # 去重子域名工作表中的内容
    unique_subdomains = list(set([row[0] for row in subdomain_ws.iter_rows(values_only=True)]))
    subdomain_ws.delete_rows(1, subdomain_ws.max_row)
    for subdomain in unique_subdomains:
        subdomain_ws.append([subdomain])        
    # 写入等号工作表
    for url in equal_urls:
        equal_ws.append([url])

    # 写入杂工作表
    for url in misc_urls:
        misc_ws.append([url])

    # 处理其他后缀的工作表
    suffix_urls = {}
    for url in urls:
        url = url.strip()

        if not url:
            continue

        if '=' in url or url.endswith('.html'):
            continue

        parsed_url = urlparse(url)
        path = parsed_url.path

        if path and '.' in path:
            parts = path.split('.')
            suffix = parts[-1].lower()

            if suffix not in suffix_urls:
                sanitized_suffix = sanitize_sheet_name(suffix)
                suffix_ws = wb.create_sheet(title=sanitized_suffix)
                suffix_urls[suffix] = suffix_ws

            suffix_ws = suffix_urls[suffix]
            suffix_ws.append([url])

    # 将子域名工作表移动到第一个位置
    wb.move_sheet(subdomain_ws, offset=-3)

    # 将等号工作表移动到第二个位置
    wb.move_sheet(equal_ws, offset=-2)

    # 将杂工作表移动到第三个位置
    wb.move_sheet(misc_ws, offset=-1)

    # 删除空白工作表
    sheets_to_delete = []
    for sheet in wb.sheetnames:
        if not sheet or sheet == 'Sheet':
            sheets_to_delete.append(sheet)

    for sheet_name in sheets_to_delete:
        del wb[sheet_name]

    # 保存 Excel 文件
    excel_file_path = os.path.join(Path.home(), 'Desktop', 'url_class.xlsx')
    wb.save(excel_file_path)

    print(f'Excel file "{excel_file_path}" created successfully.')

# 解析命令行参数
parser = argparse.ArgumentParser()
parser.add_argument('-f', '--file', help='Path to the input TXT file')
args = parser.parse_args()

# 处理传入的 TXT 文件
if args.file:
    try:
        classify_urls(args.file)
    except ValueError as e:
        print(f'Error: {str(e)}')
else:
    print('Please provide the path to the input TXT file using the -f or --file option.')
